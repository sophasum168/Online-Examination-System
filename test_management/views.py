# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from openpyxl import Workbook, load_workbook
import json, csv, logging, ast
from datetime import datetime
from django.contrib import messages
from django.shortcuts import render, redirect
from django.http.response import JsonResponse
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import render_to_string
from rest_framework.parsers import JSONParser
from rest_framework.renderers import JSONRenderer
from .models import *
from .forms import QuestionForm, OptionForm
from .serializers import QuestionSerializer, OptionSerializer
from register.models import Register
from datetime import datetime
from django.contrib.auth.decorators import login_required
# from .forms import TestUpload


# Create your views here.
# login
@login_required(login_url="/login/")

def test_list(request):
    tests = Test.objects.all()
    return render(request, 'test_list.html', {'tests' : tests})

def add_test(request):
    if request.method == 'POST':
        test_name = json.loads(request.POST.get('test_name'))
        test_type = json.loads(request.POST.get('test_type'))
        test_date = json.loads(request.POST.get('test_date'))
        test_date = datetime.strptime(test_date, '%Y-%m-%d').date()
        test_obj = Test(test_name = test_name, test_type = test_type, test_date = test_date)
        test_obj.save()
        return redirect('test')

def edit_test(request):
    if request.method == 'POST':
        row_id = json.loads(request.POST.get('row_id'))
        test_name = json.loads(request.POST.get('test_name'))
        test_type = json.loads(request.POST.get('test_type'))
        test_date = json.loads(request.POST.get('test_date'))

        test_date = datetime.strptime(test_date, '%Y-%m-%d').date()
        row_id = int(row_id)
        tests = Test.objects.all()
        
        for test_obj in tests:
            if test_obj.id == row_id:
                test_obj.test_name = test_name
                test_obj.test_type = test_type
                test_obj.test_date = test_date
                test_obj.save()
                return HttpResponseRedirect('/test/')

def delete_test(request):
    if request.is_ajax():
        selected_tests = request.POST['test_list_ids']
        selected_tests = json.loads(selected_tests)
        for i, test in enumerate(selected_tests):
            Test.objects.filter(id__in=selected_tests).delete()
        return redirect('test')

#login
@login_required(login_url="/login/")

def question_list(request):
    questions = Question.objects.all()
    questionForm = QuestionForm()
    optionForm = OptionForm()
    return render(request, 'question_list.html', {'questions' : questions, 'questionForm' : questionForm, 'optionForm' : optionForm })

def add_question(request):
    if request.method == 'POST':
        question_name = json.loads(request.POST.get('question_name'))
        question_type = json.loads(request.POST.get('question_type'))
        test_id = json.loads(request.POST.get('test_id'))
        if bool(request.FILES.get('question_img', False)) == False:
            question_obj = Question(question_name = question_name, question_type = question_type, test_id = Test.objects.get(id=test_id))
        else:
            img_option = request.FILES['question_img']
            print ("img_option",img_option)
            question_obj = Question(img_option = img_option, question_name = question_name, question_type = question_type, test_id = Test.objects.get(id=test_id))
        context = question_obj.save(request)
        
        option_rows = json.loads(request.POST.get("option_rows"))
        
        for index, option in enumerate(option_rows, start=0):
            option_obj = Option()
            option_obj.question_id = question_obj
            option_obj.option_name = option_rows[index]["option_name"]
            option_obj.answer = option_rows[index]["answer"]
            if bool(request.FILES.get('option_img'+str(index), False)) == True:
                option_obj.img_option = request.FILES['option_img'+str(index)]
            option_obj.save()
        return render(request, 'question_list.html', {'context' : context})

def edit_question(request):
    context = dict()
    if request.method == 'GET':
        row_id = request.GET.get('row_id')
        question_obj = Question()
        question = question_obj.edit_question(row_id)
        question_form = QuestionForm(initial={'question_name': question[0]['question_name'],
                                            'question_type': question[0]['question_type'],
                                            'test_id': question[0]['test_id'],
                                            'img_option': question[0]['img_option']}, auto_id=False).__str__()
        option_form = ""
        if question[1]:
            cont = dict()
            options_form = []
            for option in question[1]:
                option_form = OptionForm(initial={'option_name': option.option_name,
                                                    'answer': option.answer,
                                                    'img_option': option.img_option})
                options_form.append(option_form)
            cont.update({'options': options_form})
            option_form = render_to_string('include/question_option_row.html', cont)
        context.update({
            'question_form': question_form,
            'option_form': option_form,
        })
        return JsonResponse(context)
    elif request.method == 'POST':
        row_id = json.loads(request.POST.get('row_id'))
        test_id = json.loads(request.POST.get('test_id'))
        question_type = json.loads(request.POST.get('question_type'))
        question_name = json.loads(request.POST.get('question_name'))
        option_rows = json.loads(request.POST.get('option_rows'))
        question_img = ""

        if bool(request.FILES.get('question_img', False)) == True:
            question_img = request.FILES.get('question_img')

        option_imgs = []
        for index, option in enumerate(option_rows, start=0):
            if bool(request.FILES.get('option_img'+str(index), False)) == True:
                option_imgs.append(request.FILES['option_img'+str(index)])

        kwargs = {"test_id":test_id, "question_type":question_type, "question_name":question_name, "question_img":question_img, "option_rows":option_rows, "option_imgs":option_imgs}
        question_obj = Question()
        question = question_obj.edit_question_save(row_id, **kwargs)
        return redirect('question')

def delete_question(request):
    if request.is_ajax():
        selected_questions = request.POST['question_list_ids']
        selected_questions = json.loads(selected_questions)
        for i, test in enumerate(selected_questions):
            Question.objects.filter(id__in=selected_questions).delete()
        return redirect('question')

def import_question(request):
    if request.method == 'GET':
        with open('static/sample/import_question_sample.csv', 'rb') as csv_file:
            response = HttpResponse(csv_file.read())
            csv_file.close()
            response['content_type'] ='text/csv'
            response['Content-Disposition'] = 'attachment; filename="import_question_sample.csv"'
            return response
    elif request.method == 'POST':
        # if not GET, then proceed
        try:
            spreadsheet_file = request.FILES["spreadsheet_file"]
            if not spreadsheet_file.name.endswith('.xlsx'):
                messages.error(request,'File is not XLSX type')
                return messages
            #if file is too large, return
            if spreadsheet_file.multiple_chunks():
                messages.error(request,"Uploaded file is too big (%.2f MB)." % (spreadsheet_file.size/(1000*1000),))
                return messages

            last_test_id = None
            last_question_id = None 
            workbook = load_workbook(spreadsheet_file)
            worksheet = workbook['IQ']
            
            for row in worksheet.iter_rows(row_offset=1):
                test_id = unicode(row[0].value)
                test_name = unicode(row[1].value)
                test_type = unicode(row[2].value)
                test_date = row[3].value
                question_type = unicode(row[4].value)
                question_name = unicode(row[5].value)
                option = unicode(row[6].value)
                answer = unicode(row[7].value)
                if test_date:
                    test_date = test_date.strftime("%m/%d/%Y")
                else:
                    test_date = "None"
                test_context = {
                    "test_name": test_name,
                    "test_type": test_type,
                    "test_date": str(test_date)
                }
                question_context = {
                    "question_type": question_type,
                    "question_name": question_name,
                    "option": option,
                    "answer": answer
                }
                if test_id == "None":
                    if (test_name and test_type and test_date) == "None":
                        if (question_type and question_name) == "None":
                            #if `question` fields are empty, add option of the same question
                            question_id = last_question_id
                            Option().add_option(question_id, **question_context)
                        #if `test` fields are empty, add question of the same test
                        else:
                            test_id = last_test_id
                            question_id = Question().add_question(test_id, **question_context)
                            last_question_id = question_id
                    #if `test_id` field is empty, add new test
                    else:
                        test = Test()
                        test_id = test.add_test(**test_context)
                        question_id = Question().add_question(test_id, **question_context)
                        last_test_id = test_id
                        last_question_id = question_id
                #`test_id` field has value, check if value match any `Test` object
                else:
                    try:
                        test_obj = Test.objects.get(id = test_id)
                    except:
                        test_obj = None
                    if test_obj:
                        question_id = Question().add_question(test_id, **question_context)
                        last_test_id = test_id
                        last_question_id = question_id
                    #`test_id` field has value, not in any `Test` object
                    else:
                        test = Test()
                        test_id = test.add_test(**test_context)
                        question_id = Question().add_question(test_id, **question_context)
                        last_test_id = test_id
                        last_question_id = question_id
                        
        except Exception as e:
                messages.error(request,"Unable to upload file. "+repr(e))
            
        return redirect('question')
    return redirect('question')

@csrf_exempt
def get_question_lists(request):
    """
    API to get a list of questions
    """
    if request.method == 'GET':
        credential = ast.literal_eval(request.body)
        email = credential['email']
        password = credential['password']
        try:
            status = Register().login_authentication(email, password)
        except Exception:
            status = False
        if status is True:
            # Adapt logic to how many questions should retrieve each time
            questions = Question.objects.all()
            question_id = list()
            for question in questions:
                if question.question_type == 'QCM':
                    if Option.objects.filter(question_id = question.id):
                        question_id.append(question.id)
            options = Option.objects.filter(question_id__in = question_id)
            option_serializer = OptionSerializer(options, many = True)
            question_serializer = QuestionSerializer(questions, many = True)
            context = {
                "status": "200",
                "message": "Success",
                "question" : question_serializer.data,
                "option" : option_serializer.data,
                "email" : email,
            }
            messages.add_message(request, messages.SUCCESS, "Succesfully login!")
            return JsonResponse(context, safe=False)
        context = {
            "status": "500",
            "message": "Error"
        }
        return JsonResponse(context, safe=False)